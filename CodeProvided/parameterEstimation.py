'''
1. To do before the experiment starts:
Read README on Github.
Understand the example and how your experiment differs from that experiment. Information is given in the code on where changes need to be made, if your experiment differs are given, but in order to understand how to change the code understand the eample will help quite a lot.
Prepare the excel file (information given in the README) and name the file Experiment.xlsx, or adjust the code.
Connect your system to communicate with the code here as explained in the README.
----------------------------------------------
2. Values to specify before the experiment starts:
Values that HAVE TO BE specified before the experiment starts.
- Amount of treatments to conduct before first prediction of Gaussian process: See line 135-137.
- Multiplier start value: See line 263-264.
- Stepwidth value: See line 200-201.
- Maximum times one treatment is conducted: See line 223-224. 
Values that CAN BE specified before the experiment starts.
- Number of variables: No line in code. This is to be specified in the excel file.
- Amount of levels for each variable: No line in code. This is to be specified in the excel file.
- Prior information on objective function: See line 163-166. If no noice is specified, comment these lines out.
- Noise of observations: See line 167 - 168. If no noice is specified, estimate the best you can.
- Whilch treatments should be conducted before first prediction of Gaussian process: See line 138-139.
----------------------------------------------
3. Note:
Do not open the excel file while running the experiment / the code is wrighting to the excel file. 
This code has no function that allows to intercept the experiment and before continuation reads the current excel file content information. It is easy to add though.
This code will try to find a minimum. If an optimum is seeked, code needs to be revised.
This code is specified for experiments seeking parameter estmation.
'''

import numpy as np
import GPy
import random
import math
import openpyxl
import string
import os
import time

class AllTrials:
    def __init__(self, point):
        self.point = point

'''
This function creates a dictionary with all possible combinations of variables and levels from an excel files that specifies these combinations.
'''
def getAllTrials(sheet):
    allTrialsDictionary = {}
    for i in range(2, sheet.max_row+1):
        point = []    
        # Each combination of variables and levels should be assinged an ID that is specified in column A of the excel file. 
        iD = sheet['A'+str(i)].value
        # Each column exept column A should hold a variable and specify for each variable which level value it has in the treatment of any given row.
        for j in string.ascii_uppercase:
            if j == 'A':
                continue
            if sheet[j+str(i)].value == None:
                break
            point.append(sheet[j+str(i)].value)
        # Each entry in the dictionary is an element of the class AllTrials.
        allTrialsDictionary[iD] = AllTrials(point)
    return allTrialsDictionary

'''
This function will create a new sheet in the excel file called 'results' that contains information on the results.
Row 1 holds the name of the column and row 2 holds a description of that column.
The funtions returns the initialized sheet.
'''
def initializeSheet(data):
    sheetResults = data.create_sheet('Results')
    sheetResults['A1'] = 'sampleCount'
    sheetResults['A2'] = 'The number of trials that have been conducted so far.'
    sheetResults['B1'] = 'currentPredMinX'
    sheetResults['B2'] = 'The point in the feature space where the currently predicted minimum is located.'
    sheetResults['C1'] = 'currentPredMinY'
    sheetResults['C2'] = 'The currently predicted minmum feature value.'
    sheetResults['D1'] = 'currentPredMinSigma'
    sheetResults['D2'] = 'The variance around the currently predicted minmum feature value.'
    sheetResults['E1'] = 'multiply'
    sheetResults['E2'] = 'The value of the multiplier.'
    sheetResults['F1'] = 'variance'
    sheetResults['F2'] = 'The variance below the current minimum.'
    sheetResults['G1'] = 'featureValue'
    sheetResults['G2'] = 'The observed value of the current trial.'
    sheetResults['H1'] = 'conductionOfTrial'
    sheetResults['H2'] = 'This number indicates how many times the current trial has been conducted so far.'
    sheetResults['I1'] = 'sampledPoint'
    sheetResults['I2'] = 'The currently considered trial.'
    return sheetResults

'''
This function provides information of the next treatments that are to be conducted in the experiment. 
It returns the observed value of the feature space.
Note: This function only receives information about one trial to be conducted. If a mini experiment contains more than one trial, this peace of code has to be adjusted.
Note further: Information on the next experimental design and the observed value is written to / wrote from a text file in the current location.
'''
def conductTrial(allTrialsDictionary, i, sheetResults):
    # for each variable value of the next treatment wright it in the file, but into a newline.
    writefile = open('nextTreatment.txt', 'a')
    for i in range(len(allTrialsDictionary[i].point)):    
        string = str(allTrialsDictionary[i].point[i])    
        writefile.write(string)
        writefile.write('\n')
    writefile.close()    
    # Every 5 seconds the code checks if the file 'results.txt' excists at a given location.
    starttime=time.time()
    while True:
        somethingThereYet = os.path.isfile("results.txt")
        if somethingThereYet == True:
            break
        time.sleep(5.0 - ((time.time() - starttime) % 5.0))
    # If the file exists the information is extraced.
    with open('results.txt') as workbook:
        for line in workbook:
            y = int(line)
    # The information on the next treatment in deleted after information on the observation has arrived.
    os.remove('nextTreatment.txt')
    # Wright observed feature value into the excel file.
    sheetResults['G'+str(sheetResults.max_row)] = y
    return y

'''
This function determines how many variables are spanning the feature space. It returns a list with the same lenth as dimensions in the feature space.
'''
def getCorrectDimesion(allTrialsDictionary, trialID): 
    pointList  = []   
    for i in range(len(allTrialsDictionary[0].point)):
        pointList.append(allTrialsDictionary[trialID].point[i])
    return pointList
    
'''
This function conducts the amount of trials that are supposed to be conducted before the Gaussian Process predicts the objective function for the first time.
It returns two lists. One holding the conducted trials and the other one the observed feature values.
'''
def conductFirstTrials(allTrialsDictionary, data, sheetResults):
    XVector = []
    YVector = []
    # For the amount of trials to conducted before the Gaussian Process predicts the objective function for the first time.
    # Here: 10% of all trials. len(allTrialsDictionary) = amount of all treatments.
    for i in range(int(len(allTrialsDictionary)/10)):
        # Select a treatment at random. Or if specific treatments should be conducted at first, specify them here.
        randomInt = random.randint(0, len(allTrialsDictionary)-1)
        # In a new row, write the number of trials conducted so far in column A and the conducted point in the feature space in column I.
        sheetResults['A'+str(sheetResults.max_row+1)] = i
        sheetResults['I'+str(sheetResults.max_row)] = str(allTrialsDictionary[randomInt].point)
        # If the current trial has been conducted before, wright the correct amount in column H. Else wright 1 in column H.
        if getCorrectDimesion(allTrialsDictionary, i) in XVector and len(XVector) > 1:
            sheetResults['H'+str(sheetResults.max_row)] = sheetResults['H'+str((len(XVector) - 1 - XVector[::-1].index([allTrialsDictionary[randomInt].point[0], allTrialsDictionary[randomInt].point[1]])) + 2)].value + 1
        else:
            sheetResults['H'+str(sheetResults.max_row)] = 1
        # Add the current point and the observed feature value to the correct list. 
        XVector.append(getCorrectDimesion(allTrialsDictionary, randomInt))
        YVector.append([conductTrial(allTrialsDictionary, randomInt, sheetResults)])
    return (XVector, YVector)
    
'''
This function calculates a prediction of the feature space, based on a given input.
Input are all trials and their observations.
The function return the prediction of the deature space from [-inf; inf].
'''
def GaussianProcess(X, Y, dimension, YAverage):
    # Define the kernel for the Gaussian process. It is dependent on the dimesion of the feature space and the average value of the objective function and its max variance. Thus change variance = 100 if your estimate differs. YAverage is the average value of all y points.
    ker = GPy.kern.Matern52(dimension, ARD=True) + GPy.kern.Bias(YAverage,variance=100)
    # Calculate the Gaussian process regression and thus the prediction.
    m = GPy.models.GPRegression(X,Y,ker)
    # This variable indicates changes of y value when progressing on the x scale. log(2): y changes significantly when x changes +- 2. Smallest and largest likely variance of : 0,first value: 0.5 and 2. Thus: second value = ((log(2) - log(2)) + (log(2) - log(0.5)) / 2) / 2 = 0.1.
    m.sum.Mat52.lengthscale.set_prior(GPy.priors.LogGaussian(math.log(2), 0.1))
    # This variable indicates variance of observation from mean observation. Variance of obersavtion spans 8 unit. Thus 4 on each side --> 2 is standarddeviation. Smallest and largest likely variance of : 0,first value: 0.5 and 2. Thus: second value = ((log(2) - log(2)) + (log(2) - log(0.5)) / 2) / 2 = 0.1. *2 as we are conisdering variances here.
    m.sum.Mat52.variance.set_prior(GPy.priors.LogGaussian(math.log(2)*2, 0.52*2))
    # This variable hold information on assumend noise of observation. 1.2 is the average noise and 1.4 is the standard deviation of the noise. Calcu.ation see line 165.
    m.likelihood.variance.set_prior(GPy.priors.LogGaussian(math.log(1.2)*2, 1.4*2))
    # Get values of regression.
    m.optimize(messages=True,max_f_eval = 1000)
    return m

'''
This function caluclates how much of the variance is below the current minimum.
It returns that amount.
'''
def getVariance(allTrialsDictionary, m, y_predMin, multiply):
    sumVariance, sumVarianceBelowMin = 0, 0
    # The total variance is the variance of all treatments summed up.
    for i in range(len(allTrialsDictionary)):
        a = np.array([getCorrectDimesion(allTrialsDictionary, i)])
        pred = m.predict(a)
        sumVariance += (pred[1][0][0] * multiply * 2)
        # If the variance of a treatment is below the current minimum, take the amount of the variance that is below the current minimum and add it to the amount of variance below the current minimum.
        if pred[0][0][0] - (pred[1][0][0] * multiply) <= y_predMin:
            sumVarianceBelowMin += ((pred[1][0][0] * multiply) - (pred[0][0][0] - y_predMin))
    # Variance is the percentage of variance below the current minimum.
    variance = sumVarianceBelowMin/sumVariance
    return variance

'''
This function checks if the multiplier is still valid or if it needs to be reducted and reduces the multiplier when necessary.
It return the new / old multiplier.
'''    
def getMultiply(multiply, allTrialsDictionary, m, data, file, sheetResults, y_predMin):
    # Calculate the variance below the current minimum.
    variance = getVariance(allTrialsDictionary, m, y_predMin, multiply)        
    # If the variance below the treshold of the current multiplier, reduce the multiplier.     
    if variance < multiply / 100:
        # stepwidth indicates how much the multiplier is decreased in each step.
        stepwidth = 1        
        multiply -= stepwidth
    # Safe the current multiplier and the variance below the current minimum in excel.
    sheetResults['E'+str(sheetResults.max_row)] = multiply
    sheetResults['F'+str(sheetResults.max_row)] = variance
    return multiply

'''
This function calculates the current multiplier and defines the next tiral(s) to conduct.
It returns information on the next trial and the value of the current multiplier.
'''    
def acquisitionFunction(allTrialsDictionary, m, data, file, sheetResults, XVector, multiply):
    nextTrialInfo = None  
    minValue, y_predMin = float("inf"), float("inf")
    # Find the best new experimental deisgn among all experimental designs.
    for i in range(len(allTrialsDictionary)):
        # Get a treatment to consider.        
        a = np.array([getCorrectDimesion(allTrialsDictionary, i)])
        # Get the Gaussian process prediction for that treatment.
        pred = m.predict(a)
        # Calculate the width of the variance using the current multiplier.
        width = (pred[0][0][0] - (pred[1][0][0] * multiply))
        # find the lowest value of all treatments.
        if width < minValue and XVector.count(list(a[0])) < 6:
            minValue = width
            nextTrialInfo = i
        # find the minimum predicted value and safe the minimum value, the variance of that value and point in the feature space of that value.
        if pred[0][0][0] < y_predMin:
            y_predMin = pred[0][0][0]
            y_predMinSigma = pred[1][0][0]
            x_predMin = a[0]
    x_predMinString = str(x_predMin)
    # Add all information on the current design to the excel file.
    if nextTrialInfo != None:
        sheetResults['A'+str(sheetResults.max_row+1)] = len(XVector) - 1
        sheetResults['B'+str(sheetResults.max_row)] = x_predMinString 
        sheetResults['C'+str(sheetResults.max_row)] = y_predMin
        sheetResults['D'+str(sheetResults.max_row)] = y_predMinSigma
        sheetResults['I'+str(sheetResults.max_row)] = str(allTrialsDictionary[nextTrialInfo].point)
        # If the current trial has been conducted before, wright the correct amount in column H. Else wright 1 in column H.
        if nextTrialInfo in XVector and len(XVector) > 1:
            sheetResults['H'+str(sheetResults.max_row)] = sheetResults['H'+str((len(XVector) - 1 - XVector[::-1].index(getCorrectDimesion(allTrialsDictionary, i))) + 2)].value + 1
        else:
            sheetResults['H'+str(sheetResults.max_row)] = 1
    # Calculate the multiplier.
    multiply = getMultiply(multiply, allTrialsDictionary, m, data, file, sheetResults, y_predMin)
    return (nextTrialInfo, multiply)

'''
This function is runs ADO and naviagtes Bayesian Optimization.
It continues until ADO is done, the experiment is finished. When finished, all results are safed in the excel file.
The function returns nothing, but it stops when the stopping criterion applys or all trials have been conducted.
'''    
def adaptiveDesignOptimization(file):
    data = openpyxl.load_workbook(file)
    sheetAllTrials = data.get_sheet_by_name('AllTrials')
    allTrialsDictionary = getAllTrials(sheetAllTrials)
    sheetResults = initializeSheet(data)    
    # Conduct all trials to be conducted before Gaussian processes make a prediction for the first time.
    allVectors = conductFirstTrials(allTrialsDictionary, data, sheetResults)    
    XVector, YVector = allVectors[0], allVectors[1]
    unfinished = True
    # This value specifies the start value of the multiplier.
    multiply = 20
    data.save(file)
    # This is the part where Bayesian Optimization starts. Each loop in the while loop contains one round of ADO / update of experimental design.
    while unfinished:
        # m is the prediction created by the Gaussian process.
        YAverage = sum(YVector)/len(YVector)        
        m = GaussianProcess(np.atleast_2d(XVector), np.atleast_2d(YVector), len(XVector[0]), YAverage)
        # nextTrialInfo[0] indicates the next trials to be conducted / the next experimental design.
        nextTrialInfo = acquisitionFunction(allTrialsDictionary, m, data, file, sheetResults, XVector, multiply)
        # If no treatment is available anymore (because all trials are conducted the maximum number of times) or the multiplier is zero, return and end ADO.
        if nextTrialInfo[1] == 0 or nextTrialInfo[0] == None:
            data.save(file)            
            # When this information is printed, the experiment is over and the information can be taken from the excel file.
            # If no print information is wanted, delete this line of code.
            print('experiment finished')
            return
        # The newly gained information from the current round of conducted treatments is added to the list that hold the infomation.
        XVector.append(getCorrectDimesion(allTrialsDictionary, nextTrialInfo[0]))
        YVector.append([conductTrial(allTrialsDictionary, nextTrialInfo[0], sheetResults)])
        # After each round of conducted treatments data added to the excel file is safed.
        data.save(file)
        # After each round of conducted treatments the code will provide information on the amount of trials that have been conducted already.
        # If no print information is wanted, delete this line of code.
        print('sampleCount', len(XVector))

   
#----------------------------##### run #####---------------------------------

# If your excel file is calles differently, adjust the name here.
simulation = adaptiveDesignOptimization('Experiment.xlsx')