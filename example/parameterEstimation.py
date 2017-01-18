import numpy as np
import GPy
import random
import math
import openpyxl
import string
import time
import os

class AllTrials:
    def __init__(self, point):
        self.point = point

def getAllTrials(sheet):
    allTrialsDictionary = {}
    for i in range(2, sheet.max_row+1):
        point = []    
        iD = sheet['A'+str(i)].value
        for j in string.ascii_uppercase:
            if j == 'A':
                continue
            if sheet[j+str(i)].value == None:
                break
            point.append(sheet[j+str(i)].value)
        allTrialsDictionary[iD] = AllTrials(point)
    return allTrialsDictionary

def initializeSheet(data):
    sheetResults = data.create_sheet('Results')
    sheetResults['A1'] = 'sampleCount'
    sheetResults['A2'] = '...'
    sheetResults['B1'] = 'currentPredMinX'
    sheetResults['B2'] = '...'
    sheetResults['C1'] = 'currentPredMinY'
    sheetResults['C2'] = '...'
    sheetResults['D1'] = 'currentPredMinSigma'
    sheetResults['D2'] = '...'
    sheetResults['E1'] = 'multiply'
    sheetResults['E2'] = '...'
    sheetResults['F1'] = 'variance'
    sheetResults['F2'] = '...'
    sheetResults['G1'] = 'featureValue'
    sheetResults['G2'] = '...'
    sheetResults['H1'] = 'conductionOfTrial'
    sheetResults['H2'] = '...'
    sheetResults['I1'] = 'sampledPoint'
    sheetResults['I2'] = '...'
    return sheetResults

def conductTrial(allTrialsDictionary, i, sheetResults):
    writefile = open('nextTreatment.txt', 'a')
    for i in range(len(allTrialsDictionary[i].point)):    
        string = str(allTrialsDictionary[i].point[i])    
        writefile.write(string)
        writefile.write('\n')
    writefile.close()    
    starttime=time.time()
    nothingThereYet = True
    while True:
        somethingThereYet = os.path.isfile("results.txt")
        if somethingThereYet == True:
            break
        print('still waiting', nothingThereYet)
        time.sleep(5.0 - ((time.time() - starttime) % 5.0))
    with open('results.txt') as workbook:
        for line in workbook:
            y = int(line)
    os.remove('nextTreatment.txt')
    sheetResults['G'+str(sheetResults.max_row)] = y
    return y

def getCorrectDimesion(allTrialsDictionary, trialID): 
    pointList  = []   
    for i in range(len(allTrialsDictionary[0].point)):
        pointList.append(allTrialsDictionary[trialID].point[i])
    return pointList
    
def conductFirstTrials(allTrialsDictionary, data, sheetResults):
    XVector = []
    YVector = []
    for i in range(int(len(allTrialsDictionary)/10)):
        randomInt = random.randint(0, len(allTrialsDictionary)-1)
        sheetResults['A'+str(sheetResults.max_row+1)] = i
        sheetResults['I'+str(sheetResults.max_row)] = str(allTrialsDictionary[randomInt].point)
        if getCorrectDimesion(allTrialsDictionary, i) in XVector and len(XVector) > 1:
            sheetResults['H'+str(sheetResults.max_row)] = sheetResults['H'+str((len(XVector) - 1 - XVector[::-1].index([allTrialsDictionary[randomInt].point[0], allTrialsDictionary[randomInt].point[1]])) + 2)].value + 1
        else:
            sheetResults['H'+str(sheetResults.max_row)] = 1
        XVector.append(getCorrectDimesion(allTrialsDictionary, randomInt))
        YVector.append([conductTrial(allTrialsDictionary, randomInt, sheetResults)])
    return (XVector, YVector)
    
def GaussianProcess(X, Y):
    ker = GPy.kern.Matern52(2, ARD=True) + GPy.kern.Bias(2,variance=100)
    m = GPy.models.GPRegression(X,Y,ker)
    m.sum.Mat52.lengthscale.set_prior(GPy.priors.LogGaussian(math.log(2), 0.1))
    m.sum.Mat52.variance.set_prior(GPy.priors.LogGaussian(math.log(2.25)*2, 0.52*2))
    m.likelihood.variance.set_prior(GPy.priors.LogGaussian(math.log(1.2)*2, 1.4*2))
    m.optimize(messages=True,max_f_eval = 1000)
    return m

def getVariance(allTrialsDictionary, m, y_predMin, multiply):
    sumVariance, sumVarianceBelowMin = 0, 0
    for i in range(len(allTrialsDictionary)):
        a = np.array([getCorrectDimesion(allTrialsDictionary, i)])
        pred = m.predict(a)
        sumVariance += (pred[1][0][0] * multiply * 2)
        if pred[0][0][0] - (pred[1][0][0] * multiply) <= y_predMin:
            sumVarianceBelowMin += ((pred[1][0][0] * multiply) - (pred[0][0][0] - y_predMin))
    variance = sumVarianceBelowMin/sumVariance
    return variance
    
def getMultiply(multiply, allTrialsDictionary, m, data, file, sheetResults, y_predMin):
    variance = getVariance(allTrialsDictionary, m, y_predMin, multiply)        
    if variance < multiply / 100:
        multiply -= 1
    sheetResults['E'+str(sheetResults.max_row)] = multiply
    sheetResults['F'+str(sheetResults.max_row)] = variance
    return multiply
    
def acquisitionFunction(allTrialsDictionary, m, data, file, sheetResults, XVector, multiply):
    nextTrialInfo = None  
    minValue, y_predMin = float("inf"), float("inf")
    for i in range(len(allTrialsDictionary)):
        a = np.array([getCorrectDimesion(allTrialsDictionary, i)])
        pred = m.predict(a)
        width = (pred[0][0][0] - (pred[1][0][0] * multiply))
        if width < minValue and XVector.count(list(a[0])) < 6:
            minValue = width
            nextTrialInfo = i
        if pred[0][0][0] < y_predMin:
            y_predMin = pred[0][0][0]
            y_predMinSigma = pred[1][0][0]
            x_predMin = a[0]
    x_predMinString = str(x_predMin)
    if nextTrialInfo != None:
        sheetResults['A'+str(sheetResults.max_row+1)] = len(XVector) - 1
        sheetResults['B'+str(sheetResults.max_row)] = x_predMinString 
        sheetResults['C'+str(sheetResults.max_row)] = y_predMin
        sheetResults['D'+str(sheetResults.max_row)] = y_predMinSigma
        sheetResults['I'+str(sheetResults.max_row)] = str(allTrialsDictionary[nextTrialInfo].point)
        if nextTrialInfo in XVector and len(XVector) > 1:
            sheetResults['H'+str(sheetResults.max_row)] = sheetResults['H'+str((len(XVector) - 1 - XVector[::-1].index(getCorrectDimesion(allTrialsDictionary, i))) + 2)].value + 1
        else:
            sheetResults['H'+str(sheetResults.max_row)] = 1
    multiply = getMultiply(multiply, allTrialsDictionary, m, data, file, sheetResults, y_predMin)
    return (nextTrialInfo, multiply)
    
def adaptiveDesignOptimization(file):
    data = openpyxl.load_workbook(file)
    sheetAllTrials = data.get_sheet_by_name('AllTrials')
    allTrialsDictionary = getAllTrials(sheetAllTrials)
    sheetResults = initializeSheet(data)    
    allVectors = conductFirstTrials(allTrialsDictionary, data, sheetResults)
    XVector, YVector = allVectors[0], allVectors[1]
    unfinished = True
    multiply = 20
    data.save(file)
    while unfinished:
        m = GaussianProcess(np.atleast_2d(XVector), np.atleast_2d(YVector))
        nextTrialInfo = acquisitionFunction(allTrialsDictionary, m, data, file, sheetResults, XVector, multiply)
        if nextTrialInfo[1] == 0 or nextTrialInfo[0] == None:
            data.save(file)            
            return
        XVector.append(getCorrectDimesion(allTrialsDictionary, nextTrialInfo[0]))
        YVector.append([conductTrial(allTrialsDictionary, nextTrialInfo[0], sheetResults)])
        data.save(file)
        print('sampleCount', len(XVector))
        
#----------------------------##### run #####---------------------------------

simulation = adaptiveDesignOptimization('Experiment.xlsx')

print('experiment finished')